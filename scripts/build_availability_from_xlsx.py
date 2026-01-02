import json
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = Path(r"C:\Users\RicardoMartinezH\Downloads\OPP_ Portafolio C1 2026.xlsx")
OUTPUT_PATH = BASE_DIR / "scripts" / "availability_payload.json"


def normalize_text(value: str) -> str:
    return (
        str(value or "")
        .lower()
        .encode("ascii", "ignore")
        .decode("ascii")
        .strip()
    )


def normalize_program_key(value: str) -> str:
    normalized = normalize_text(value)
    for prefix in (
        "licenciatura en ",
        "maestria en ",
        "ingenieria en ",
        "ingenieria ",
        "licenciatura ",
        "maestria ",
    ):
        if normalized.startswith(prefix):
            normalized = normalized[len(prefix) :]
            break
    return " ".join(normalized.split()).strip()


def to_title_case(value: str) -> str:
    return " ".join(word[:1].upper() + word[1:].lower() for word in str(value).split())


def parse_availability(raw) -> bool:
    if raw is True:
        return True
    if raw is False:
        return False
    text = str(raw or "").strip()
    normalized = normalize_text(text)
    if not normalized:
        return False
    if any(char in text for char in ("\u2713", "\u2714", "\u2705")):
        return True
    if normalized in ("si", "true", "1", "disponible", "activo", "verdadero"):
        return True
    return False


def extract_links(sheet):
    links_by_row = {}
    links_by_cell = {}
    hidden_rows = set()
    for row in sheet.iter_rows():
        row_idx = row[0].row - 1
        if sheet.row_dimensions[row[0].row].hidden:
            hidden_rows.add(row_idx)
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                links_by_cell[(row_idx, cell.column - 1)] = cell.hyperlink.target
                links_by_row.setdefault(row_idx, cell.hyperlink.target)
    return links_by_row, links_by_cell, hidden_rows


def build_online_availability(rows, sheet_name, links_by_row, links_by_cell, hidden_rows):
    header_matches = []
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            normalized = normalize_text(cell)
            if not normalized or "online" not in normalized:
                continue
            if "licenciatura" in normalized:
                header_matches.append((r_idx, c_idx, "licenciatura online"))
            elif "posgrados" in normalized or "maestria" in normalized:
                header_matches.append((r_idx, c_idx, "posgrados online"))

    header_matches.sort()
    lic_headers = [h for h in header_matches if h[2] == "licenciatura online"]
    pos_headers = [h for h in header_matches if h[2] == "posgrados online"]
    pos_start = min([h[0] for h in pos_headers], default=None)

    entries = []

    def resolve_link(r_idx, c_idx):
        return links_by_cell.get((r_idx, c_idx)) or links_by_row.get(r_idx, "")

    def parse_headers(headers, end_override=None):
        for idx, (row_idx, col_idx, label) in enumerate(headers):
            next_row = None
            for h in header_matches:
                if h[0] > row_idx:
                    next_row = h[0]
                    break
            end_row = next_row if next_row is not None else len(rows)
            if end_override is not None:
                end_row = min(end_row, end_override)
            for r in range(row_idx + 1, end_row):
                if r in hidden_rows:
                    continue
                cell_value = rows[r][col_idx] if col_idx < len(rows[r]) else ""
                programa = to_title_case(str(cell_value or "").strip())
                if not programa:
                    continue
                normalized = normalize_text(programa)
                if "online" in normalized and ("licenciatura" in normalized or "posgrados" in normalized):
                    continue
                if normalized in ("programa", "programas"):
                    continue
                entries.append(
                    {
                        "id": f"sheet-{sheet_name}-{label}-{r}-{col_idx}-online",
                        "plantel": sheet_name,
                        "programa": programa,
                        "modalidad": "online",
                        "horario": "",
                        "planUrl": resolve_link(r, col_idx),
                        "activo": True,
                    }
                )

    parse_headers(lic_headers, pos_start)
    parse_headers(pos_headers)
    return entries


def build_sheet_availability(rows, sheet_name, links_by_row, links_by_cell, hidden_rows):
    normalized_rows = [[str(cell or "") for cell in row] for row in rows]
    header_idx = -1
    for idx, row in enumerate(normalized_rows):
        has_c1 = any("c1" in normalize_text(cell) for cell in row)
        has_2026 = any("2026" in normalize_text(cell) for cell in row)
        if has_c1 and has_2026:
            header_idx = idx
            break
    if header_idx < 0:
        return []

    year_idx = header_idx
    for idx in range(header_idx, min(header_idx + 6, len(normalized_rows))):
        if any(normalize_text(cell) == "2026" for cell in normalized_rows[idx]):
            year_idx = idx
            break

    modalidad_idx = -1
    for idx in range(year_idx, min(year_idx + 4, len(normalized_rows))):
        if any("escolarizado" in normalize_text(cell) or "ejecutivo" in normalize_text(cell) for cell in normalized_rows[idx]):
            modalidad_idx = idx
            break
    if modalidad_idx < 0:
        modalidad_idx = min(year_idx + 1, len(normalized_rows) - 1)

    modalidad_row = normalized_rows[modalidad_idx]
    escolarizado_cols = [i for i, cell in enumerate(modalidad_row) if "escolarizado" in normalize_text(cell)]
    ejecutivo_cols = [i for i, cell in enumerate(modalidad_row) if "ejecutivo" in normalize_text(cell)]

    header_row = normalized_rows[header_idx]
    horarios_header_col = next((i for i, cell in enumerate(header_row) if normalize_text(cell) == "horarios"), -1)
    availability_escolarizado_cols = [c for c in escolarizado_cols if c < horarios_header_col] if horarios_header_col >= 0 else escolarizado_cols
    availability_ejecutivo_cols = [c for c in ejecutivo_cols if c < horarios_header_col] if horarios_header_col >= 0 else ejecutivo_cols

    escolarizado_col = availability_escolarizado_cols[0] if availability_escolarizado_cols else (escolarizado_cols[0] if escolarizado_cols else -1)
    ejecutivo_col = availability_ejecutivo_cols[0] if availability_ejecutivo_cols else (ejecutivo_cols[0] if ejecutivo_cols else -1)
    if escolarizado_col < 0:
        escolarizado_col = 2
    if ejecutivo_col < 0:
        ejecutivo_col = 3

    schedule_escolarizado_col = next((c for c in escolarizado_cols if c > horarios_header_col), -1) if horarios_header_col >= 0 else -1
    schedule_ejecutivo_col = next((c for c in ejecutivo_cols if c > horarios_header_col), -1) if horarios_header_col >= 0 else -1
    schedule_escolarizado_fallback = schedule_escolarizado_col if schedule_escolarizado_col >= 0 else 7
    schedule_ejecutivo_fallback = schedule_ejecutivo_col if schedule_ejecutivo_col >= 0 else 8

    horarios_idx = next((i for i, row in enumerate(normalized_rows) if any(normalize_text(cell) == "horarios" for cell in row)), -1)
    end_idx = horarios_idx if horarios_idx > modalidad_idx else len(normalized_rows)

    entries = []
    for offset, row in enumerate(normalized_rows[modalidad_idx + 1 : end_idx]):
        real_idx = modalidad_idx + 1 + offset
        if real_idx in hidden_rows:
            continue
        programa = (row[1] if len(row) > 1 and row[1].strip() else (row[0] if row else "")).strip()
        if not programa:
            continue
        programa_norm = normalize_text(programa)
        if programa_norm in ("modular", "longitudinal", "programa", "programas"):
            continue
        escolarizado_activo = parse_availability(row[escolarizado_col] if escolarizado_col >= 0 else "")
        ejecutivo_activo = parse_availability(row[ejecutivo_col] if ejecutivo_col >= 0 else "")
        if not escolarizado_activo and not ejecutivo_activo:
            continue

        program_col = 1 if len(row) > 1 and row[1].strip() else 0
        plan_url = links_by_cell.get((real_idx, program_col)) or links_by_row.get(real_idx, "")

        if escolarizado_activo:
            horario = str(row[schedule_escolarizado_fallback] if schedule_escolarizado_fallback >= 0 else "").strip()
            entries.append(
                {
                    "id": f"sheet-{sheet_name}-{offset}-presencial",
                    "plantel": sheet_name,
                    "programa": to_title_case(programa),
                    "modalidad": "presencial",
                    "horario": horario,
                    "planUrl": plan_url,
                    "activo": True,
                }
            )
        if ejecutivo_activo:
            horario = str(row[schedule_ejecutivo_fallback] if schedule_ejecutivo_fallback >= 0 else "").strip()
            entries.append(
                {
                    "id": f"sheet-{sheet_name}-{offset}-mixta",
                    "plantel": sheet_name,
                    "programa": to_title_case(programa),
                    "modalidad": "mixta",
                    "horario": horario,
                    "planUrl": plan_url,
                    "activo": True,
                }
            )
    return entries


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    availability = []
    debug = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=False)]
        links_by_row, links_by_cell, hidden_rows = extract_links(sheet)

        if normalize_text(sheet_name) == "oferta general":
            continue
        if "online" in normalize_text(sheet_name):
            entries = build_online_availability(rows, sheet_name, links_by_row, links_by_cell, hidden_rows)
        else:
            entries = build_sheet_availability(rows, sheet_name, links_by_row, links_by_cell, hidden_rows)
        availability.extend(entries)
        debug.append({"plantel": sheet_name, "entries": len(entries)})

    # plan URL fallback for online
    plan_url_by_program = {}
    for entry in availability:
        if entry["modalidad"] == "online":
            continue
        key = normalize_program_key(entry["programa"])
        if entry.get("planUrl") and key and key not in plan_url_by_program:
            plan_url_by_program[key] = entry["planUrl"]

    for entry in availability:
        if entry["modalidad"] != "online":
            continue
        if entry.get("planUrl"):
            continue
        key = normalize_program_key(entry["programa"])
        if key in plan_url_by_program:
            entry["planUrl"] = plan_url_by_program[key]

    # add missing online programs from allowlist
    allowlist = [
        "Licenciatura en Administración de Empresas",
        "Licenciatura en Administración de Empresas Turísticas",
        "Licenciatura en Administración de Tecnologías de la Información",
        "Licenciatura en Contaduría Pública",
        "Licenciatura en Ciencias de la Comunicación",
        "Licenciatura en Comercio Internacional",
        "Licenciatura en Mercadotecnia",
        "Licenciatura en Derecho",
        "Licenciatura en Diseño Gráfico",
        "Licenciatura en Arquitectura",
        "Licenciatura en Pedagogía",
        "Ingeniería Industrial y de Sistemas",
        "Ingeniería en Manufactura y Robótica",
        "Ingeniería en Sistemas Computacionales",
        "Licenciatura en Relaciones Internacionales",
        "Licenciatura en Negocios Internacionales",
        "Licenciatura en Economía y Finanzas",
        "Licenciatura en Administración Financiera",
        "Licenciatura en Administración de Recursos Humanos",
        "Ingeniería Industrial y Administración",
        "Ingeniería en Software y Redes",
        "Ingeniería en Logística",
        "Licenciatura en Seguridad Pública",
        "Licenciatura en Criminología",
        "Maestría en Administración de Negocios",
        "Maestría en Administración Financiera",
        "Maestría en Mercadotecnia",
        "Maestría en Gestión de Talento Humano",
        "Maestría en Gestión de Proyectos",
        "Maestría en Derecho Constitucional y Amparo",
        "Maestría en Derecho Corporativo",
        "Maestría en Derecho Fiscal y Administrativo",
        "Maestría en Derecho Laboral",
        "Maestría en Derecho Procesal",
        "Maestría en Derecho y Juicios Orales",
        "Maestría en Educación y Docencia",
        "Maestría en Gestión Educativa",
        "Maestría en Administración de Servicios de Salud",
        "Maestría en Administración de Negocios y Mercadotecnia",
        "Maestría en Finanzas",
        "Maestría en Administración Pública",
        "Maestría en Diseño Digital",
        "Maestría en Diseño Sostenible y Arquitectura Verde",
        "Maestría en Diseño Estratégico e Innovación",
        "Maestría en Robótica y Automatización",
        "Maestría en Inteligencia Artificial",
        "Maestría en Energías Renovables",
        "Maestría en Interacción y Experiencia del Usuario",
        "Maestría en Logística y Cadena de Suministro",
    ]
    online_keys = {normalize_program_key(entry["programa"]) for entry in availability if entry["modalidad"] == "online"}
    for programa in allowlist:
        key = normalize_program_key(programa)
        if key in online_keys:
            continue
        availability.append(
            {
                "id": f"online-allowlist-{key}",
                "plantel": "Online",
                "programa": to_title_case(programa),
                "modalidad": "online",
                "horario": "",
                "planUrl": plan_url_by_program.get(key, ""),
                "activo": True,
            }
        )

    deduped = {}
    non_online = []
    for entry in availability:
        if entry["modalidad"] != "online":
            non_online.append(entry)
            continue
        key = f"{normalize_program_key(entry['programa'])}::online::{normalize_text(entry['plantel'])}"
        current = deduped.get(key)
        if not current:
            deduped[key] = entry
            continue
        if not current.get("planUrl") and entry.get("planUrl"):
            deduped[key] = entry
            continue
        if not current.get("horario") and entry.get("horario"):
            deduped[key] = entry
    payload = {"availability": non_online + list(deduped.values()), "debug": debug}
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(availability)} entries to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
